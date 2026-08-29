#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
from pathlib import Path

from protocol_v54 import read_tsv, write_json

REQUIRED = {"source_record_id", "report_id", "study_id", "main_source_path", "main_source_sha256"}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""): h.update(block)
    return h.hexdigest()


def read_index(path: Path, sheet: str) -> tuple[list[str], list[dict[str, str]]]:
    if path.suffix.lower() in {".tsv", ".txt"}: return read_tsv(path)
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    if sheet not in wb.sheetnames: raise ValueError(f"missing sheet {sheet}")
    ws = wb[sheet]; rows = ws.iter_rows(values_only=True)
    first = next(rows)
    if first and first[0] == sheet: next(rows)
    headers = [str(x) for x in next(rows)]
    return headers, [{headers[i]: "" if v is None else str(v) for i, v in enumerate(row)} for row in rows if any(v is not None for v in row)]


def validate(path: Path, sheet: str, expected_sources: int | None, expected_reports: int | None,
             verify_files: bool, compat_v45_placeholders: bool = False) -> dict:
    errors: list[str] = []; warnings: list[str] = []
    try: headers, rows = read_index(path, sheet)
    except Exception as exc: return {"protocol": "SOURCE_IDENTITY_V1/v5.4", "pass": False, "errors": [str(exc)], "warnings": []}
    missing = sorted(REQUIRED - set(headers))
    if missing: errors.append(f"missing columns: {missing}")
    seen_sources = set(); report_to_study = {}
    for n, row in enumerate(rows, 2):
        src, report, study = (row.get(k, "").strip() for k in ("source_record_id", "report_id", "study_id"))
        if not src or src in seen_sources:
            if compat_v45_placeholders and src == "NR_ORIGINAL_ID" and report and study:
                warnings.append(f"row {n}: legacy NR_ORIGINAL_ID requires canonical_source_record_id derived from report/DOI")
            else:
                errors.append(f"row {n}: blank/duplicate source_record_id {src!r}")
        seen_sources.add(src)
        if not report or not study: errors.append(f"row {n}: blank report/study")
        if report in report_to_study and report_to_study[report] != study:
            errors.append(f"row {n}: report maps to multiple studies: {report}")
        report_to_study[report] = study
        if verify_files:
            p = Path(row.get("main_source_path", ""))
            if not p.exists(): errors.append(f"row {n}: missing source {p}")
            elif sha256(p).lower() != row.get("main_source_sha256", "").lower(): errors.append(f"row {n}: main hash mismatch")
            sp = row.get("supplement_paths", "").strip()
            if sp:
                spp = Path(sp)
                if not spp.exists(): errors.append(f"row {n}: missing supplement {spp}")
                elif sha256(spp).lower() != row.get("supplement_sha256", "").lower(): errors.append(f"row {n}: supplement hash mismatch")
    if expected_sources is not None and len(rows) != expected_sources: errors.append(f"expected {expected_sources} source records, observed {len(rows)}")
    if expected_reports is not None and len(report_to_study) != expected_reports: errors.append(f"expected {expected_reports} reports, observed {len(report_to_study)}")
    return {"protocol": "SOURCE_IDENTITY_V1/v5.4", "pass": not errors, "errors": errors, "warnings": warnings,
            "counts": {"source_records": len(rows), "reports": len(report_to_study), "studies": len(set(report_to_study.values()))},
            "identity_rule": "Exact source index mapping; row/file order is never used."}


def main() -> int:
    p=argparse.ArgumentParser(); p.add_argument("source_index",type=Path); p.add_argument("--sheet",default="01_Source_Report_Index")
    p.add_argument("--expected-sources",type=int); p.add_argument("--expected-reports",type=int); p.add_argument("--verify-files",action="store_true"); p.add_argument("--compat-v45-placeholders",action="store_true"); p.add_argument("--out",type=Path)
    a=p.parse_args(); result=validate(a.source_index,a.sheet,a.expected_sources,a.expected_reports,a.verify_files,a.compat_v45_placeholders); write_json(result,a.out); return 0 if result["pass"] else 1


if __name__=="__main__": raise SystemExit(main())
