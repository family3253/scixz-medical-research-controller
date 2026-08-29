#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook

from recover_schema_drift_v57 import clean, read_tsv


def sheet_name(entity_type: str) -> str:
    return (re.sub(r"[\\/*?:\[\]]", "_", entity_type) + "_recovered")[:31]


def excel_value(value: object) -> str:
    text = clean(value)
    if len(text) > 32767:
        return text[:32740] + "...[TRUNCATED_IN_XLSX]"
    return text


def write_table(ws, headers: list[str], rows: list[dict[str, str]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([excel_value(row.get(field)) for field in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_workbook(recovery_root: Path, output: Path) -> dict:
    manifest_path = recovery_root / "recovery_manifest.json"
    coverage_path = recovery_root / "companion_coverage.tsv"
    if not manifest_path.is_file() or not coverage_path.is_file():
        raise ValueError("recovery manifest and companion coverage are required")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    tables = list(manifest.get("selected_tables") or [])
    if not tables:
        raise ValueError("manifest selected_tables is empty")
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists():
        raise ValueError(f"refusing to overwrite existing workbook: {output}")

    wb = Workbook(write_only=False)
    wb.remove(wb.active)
    sheet_map: list[dict[str, str]] = []
    expected_rows: dict[str, int] = {}
    for entity_type in tables:
        view_path = recovery_root / "views" / f"{entity_type}_recovered.tsv"
        if not view_path.is_file():
            raise ValueError(f"missing recovered view: {view_path.name}")
        headers, rows = read_tsv(view_path)
        name = sheet_name(entity_type)
        write_table(wb.create_sheet(name), headers, rows)
        expected_rows[name] = len(rows)
        sheet_map.append({
            "entity_type": entity_type,
            "sheet_name": name,
            "source_tsv": str(view_path),
            "source_sha256": hashlib.sha256(view_path.read_bytes()).hexdigest(),
            "data_rows": str(len(rows)),
        })

    coverage_headers, coverage_rows = read_tsv(coverage_path)
    write_table(wb.create_sheet("companion_coverage"), coverage_headers, coverage_rows)
    expected_rows["companion_coverage"] = len(coverage_rows)
    map_headers = ["entity_type", "sheet_name", "source_tsv", "source_sha256", "data_rows"]
    write_table(wb.create_sheet("__sheet_map"), map_headers, sheet_map)
    manifest_ws = wb.create_sheet("__manifest")
    manifest_ws.append(["recovery_manifest_json"])
    manifest_ws.append([json.dumps(manifest, ensure_ascii=False, sort_keys=True)])
    wb.save(output)
    return validate_workbook(recovery_root, output, expected_rows)


def validate_workbook(recovery_root: Path, workbook: Path,
                      expected_rows: dict[str, int] | None = None) -> dict:
    errors: list[str] = []
    manifest = json.loads((recovery_root / "recovery_manifest.json").read_text(encoding="utf-8"))
    _, coverage = read_tsv(recovery_root / "companion_coverage.tsv")
    wb = load_workbook(workbook, read_only=True, data_only=False)
    view_index: dict[str, dict[tuple[str, str, str], dict[str, object]]] = {}
    row_counts: dict[str, int] = {}
    for entity_type in manifest.get("selected_tables") or []:
        name = sheet_name(entity_type)
        if name not in wb.sheetnames:
            errors.append(f"missing workbook sheet {name}")
            continue
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        headers = [clean(x) for x in next(rows, ())]
        table: dict[tuple[str, str, str], dict[str, object]] = {}
        count = 0
        for values in rows:
            count += 1
            row = dict(zip(headers, values))
            key = tuple(clean(row.get(k)) for k in ("report_id", "study_id", "entity_id"))
            if not all(key) or key in table:
                errors.append(f"{name}: blank or duplicate entity key {key}")
            table[key] = row
        view_index[entity_type] = table
        row_counts[name] = count
    if "companion_coverage" not in wb.sheetnames:
        errors.append("missing workbook sheet companion_coverage")
    else:
        row_counts["companion_coverage"] = max(wb["companion_coverage"].max_row - 1, 0)

    if expected_rows:
        for name, expected in expected_rows.items():
            if row_counts.get(name) != expected:
                errors.append(f"{name}: expected {expected} rows, found {row_counts.get(name)}")

    checked = 0
    for row_number, claim in enumerate(coverage, 2):
        if clean(claim.get("visible_01")) != "1":
            errors.append(f"coverage row {row_number} is not marked visible")
            continue
        entity_type = clean(claim.get("entity_type"))
        key = tuple(clean(claim.get(k)) for k in ("report_id", "study_id", "entity_id"))
        field = clean(claim.get("field_name"))
        expected = clean(claim.get("result_value"))
        workbook_row = view_index.get(entity_type, {}).get(key)
        if workbook_row is None:
            errors.append(f"coverage row {row_number} has no workbook entity {entity_type}:{key}")
            continue
        actual = clean(workbook_row.get(field))
        status = clean(workbook_row.get(field + "__status")).upper()
        if actual != expected or status != "OBSERVED":
            errors.append(
                f"coverage row {row_number} not materialized in workbook: expected {expected!r}/OBSERVED, found {actual!r}/{status!r}"
            )
        checked += 1
    wb.close()
    result = {
        "protocol": "v5.7-workbook-export",
        "pass": not errors,
        "workbook": str(workbook),
        "workbook_sha256": hashlib.sha256(workbook.read_bytes()).hexdigest(),
        "counts": {"companion_values_checked": checked, "sheets": len(row_counts), "rows": row_counts},
        "errors": errors,
    }
    qa_path = workbook.with_suffix(workbook.suffix + ".qa.json")
    qa_path.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Export recovered MDR-GNB views and prove companion values are visible in XLSX")
    parser.add_argument("--recovery-root", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    result = export_workbook(args.recovery_root, args.output)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["pass"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
