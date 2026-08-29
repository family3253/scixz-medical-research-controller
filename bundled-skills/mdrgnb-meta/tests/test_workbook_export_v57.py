from __future__ import annotations

import csv
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from export_recovered_workbook_v57 import export_workbook, validate_workbook
from recover_schema_drift_v57 import recover


def write_tsv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", lineterminator="\n")
        writer.writeheader(); writer.writerows(rows)


class WorkbookExportV57Tests(unittest.TestCase):
    def test_export_roundtrip_blocks_false_nr_in_xlsx(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td); canonical = root / "canonical"; recovery = root / "recovery"; canonical.mkdir()
            write_tsv(canonical / "performance.tsv",
                      ["performance_id", "study_id", "report_id", "metric_code"],
                      [{"performance_id": "P1", "study_id": "S1", "report_id": "R1", "metric_code": "AUC"}])
            write_tsv(canonical / "performance_values.tsv",
                      ["performance_id", "stu", "metric_code", "estimate"],
                      [{"performance_id": "P1", "stu": "S1", "metric_code": "AUC", "estimate": "0.91"}])
            recover(canonical, [], recovery, selected_tables={"performance"},
                    required_companions={"performance_values.tsv"})
            workbook = root / "recovered.xlsx"
            result = export_workbook(recovery, workbook)
            self.assertTrue(result["pass"], result)
            self.assertEqual(result["counts"]["companion_values_checked"], 4)

            wb = load_workbook(workbook)
            ws = wb["performance_recovered"]
            headers = {cell.value: cell.column for cell in ws[1]}
            ws.cell(2, headers["estimate"], "NR")
            tampered = root / "tampered.xlsx"
            wb.save(tampered); wb.close()
            tampered_result = validate_workbook(recovery, tampered)
            self.assertFalse(tampered_result["pass"])
            self.assertTrue(any("not materialized in workbook" in e for e in tampered_result["errors"]))


if __name__ == "__main__":
    unittest.main()
