from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import bibtexparser
from bibtexparser.bwriter import BibTexWriter
from habanero import Crossref
import httpx
from openpyxl import load_workbook


DEFAULT_COLLECTION_NAME = "cpu"
DEFAULT_COLLECTION_KEY = "UDC49MTG"
DEFAULT_STATUS_PREFIX = "✓"
DEFAULT_CHECKPOINT_EVERY = 20
DEFAULT_TIMEOUT = 20
DEFAULT_RETRIES = 1
DOI_RE = re.compile(r"(10\.\d{4,9}/[-._;()/:A-Z0-9]+)", re.I)


@dataclass
class ReviewedRow:
    row_number: int
    status: str
    identifier: str
    reference_text: str
    review_note: str


@dataclass
class ImportResult:
    row_number: int
    identifier: str
    reference_text: str
    source_type: str
    action: str
    zotero_key: str = ""
    title: str = ""
    doi: str = ""
    url: str = ""
    reason: str = ""


@dataclass
class LocalWriteProbe:
    supported: bool
    checked_at: str
    collection_key: str
    collection_found: bool
    created_key: str = ""
    cleaned_up: bool = False
    reason: str = ""
    raw_result: str = ""


def make_datacite_citekey(title: str, authors: list[str], year: str | int | None) -> str:
    family = re.sub(r"[^a-z0-9]+", "", authors[0].split()[-1].lower()) if authors else "ref"
    words = [re.sub(r"[^a-z0-9]+", "", word.lower()) for word in title.split()]
    words = [word for word in words if word]
    short = "".join(words[:3])[:18] or "item"
    return f"{family}{short}{year or 'nd'}"


def extract_doi(value: str) -> str:
    match = DOI_RE.search(value or "")
    if not match:
        return ""
    return match.group(1).rstrip(".,);]")


def normalize_source_type(identifier: str) -> str:
    if not identifier:
        return "text"
    if DOI_RE.search(identifier):
        return "doi"
    if identifier.lower().startswith(("http://", "https://")):
        return "url"
    return "text"


def sha1_file(path: Path) -> str:
    digest = hashlib.sha1()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def describe_exception(exc: Exception) -> str:
    parts = [type(exc).__name__]
    response = getattr(exc, "response", None)
    if response is not None:
        status_code = getattr(response, "status_code", None)
        if status_code is not None:
            parts.append(f"status={status_code}")
        text = getattr(response, "text", "")
        if text:
            parts.append(text.strip())
    message = str(exc).strip()
    if message:
        parts.append(message)
    return " | ".join(part for part in parts if part)


def write_json_atomic(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(path)


def load_json(path: Path, default: dict[str, Any] | None = None) -> dict[str, Any]:
    if not path.exists():
        return default.copy() if default else {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default.copy() if default else {}


def load_reviewed_rows(path: Path, sheet_name: str | None = None, status_prefix: str = DEFAULT_STATUS_PREFIX) -> list[ReviewedRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    headers = [str(cell).strip() if cell is not None else "" for cell in next(sheet.iter_rows(values_only=True))]
    index = {name: i for i, name in enumerate(headers)}

    required = ["状态", "DOI/URL", "参考文献文本(前120字)", "审核说明"]
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    rows: list[ReviewedRow] = []
    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=2):
        status = str(values[index["状态"]] or "").strip()
        if not status.startswith(status_prefix):
            continue

        identifier = str(values[index["DOI/URL"]] or "").strip()
        reference_text = str(values[index["参考文献文本(前120字)"]] or "").strip()
        review_note = str(values[index["审核说明"]] or "").strip()

        if not identifier:
            identifier = extract_doi(reference_text)
        if not identifier and not reference_text:
            continue

        rows.append(
            ReviewedRow(
                row_number=row_number,
                status=status,
                identifier=identifier,
                reference_text=reference_text,
                review_note=review_note,
            )
        )

    return rows


def datacite_to_message(doi: str, timeout: int) -> dict[str, Any] | None:
    try:
        response = httpx.get(f"https://api.datacite.org/dois/{doi}", timeout=timeout)
        response.raise_for_status()
        payload = response.json()
    except Exception:
        return None

    attributes = payload.get("data", {}).get("attributes", {})
    titles = attributes.get("titles", [])
    creators = attributes.get("creators", [])
    title = titles[0].get("title", "").strip() if titles else ""
    creator_names: list[str] = []
    for creator in creators:
        name = creator.get("name") or " ".join(
            part for part in [creator.get("givenName", "").strip(), creator.get("familyName", "").strip()] if part
        )
        if name:
            creator_names.append(name)
    creator_dicts: list[dict[str, str]] = []
    for name in creator_names:
        parts = name.split()
        if not parts:
            continue
        creator_dicts.append({"given": " ".join(parts[:-1]), "family": parts[-1]})

    issued = {}
    publication_year = attributes.get("publicationYear")
    if publication_year:
        issued = {"date-parts": [[int(publication_year)]]}

    return {
        "title": [title] if title else [],
        "DOI": doi,
        "URL": f"https://doi.org/{doi}",
        "container-title": [attributes.get("publisher", "")],
        "author": creator_dicts,
        "issued": issued,
    }


def fetch_crossref_payload(
    identifier: str,
    fallback_reference: str,
    cr: Crossref,
    *,
    metadata_cache: dict[str, dict[str, Any] | None],
    timeout: int,
    retries: int,
) -> tuple[dict[str, Any] | None, str]:
    doi = extract_doi(identifier or fallback_reference)
    if doi:
        cache_key = f"doi:{doi.lower()}"
        if cache_key in metadata_cache:
            cached = metadata_cache[cache_key]
            return (cached or {}).get("message"), doi

        last_exc: Exception | None = None
        for _ in range(max(1, retries + 1)):
            try:
                message = cr.works(ids=doi)["message"]
                metadata_cache[cache_key] = {"message": message}
                return message, doi
            except Exception as exc:
                last_exc = exc

        datacite = datacite_to_message(doi, timeout=timeout)
        if datacite is not None:
            metadata_cache[cache_key] = {"message": datacite}
            return datacite, doi

        metadata_cache[cache_key] = None
        if last_exc:
            raise last_exc
        return None, doi

    if not fallback_reference:
        return None, ""

    ref_key = hashlib.sha1(fallback_reference.encode("utf-8")).hexdigest()
    cache_key = f"bib:{ref_key}"
    if cache_key in metadata_cache:
        cached = metadata_cache[cache_key]
        if cached:
            return cached.get("message"), cached.get("doi", "")
        return None, ""

    last_exc = None
    for _ in range(max(1, retries + 1)):
        try:
            response = cr.works(query_bibliographic=fallback_reference, limit=1)
            items = response.get("message", {}).get("items", [])
            if items:
                result_doi = items[0].get("DOI", "")
                metadata_cache[cache_key] = {"message": items[0], "doi": result_doi}
                return items[0], result_doi
            metadata_cache[cache_key] = None
            return None, ""
        except Exception as exc:
            last_exc = exc

    metadata_cache[cache_key] = None
    if last_exc:
        raise last_exc
    return None, ""


def payload_to_bib_entry(message: dict[str, Any], fallback_reference: str, row_number: int) -> dict[str, str]:
    title_list = message.get("title", [])
    title = title_list[0].strip() if title_list else fallback_reference
    doi = message.get("DOI", "")
    url = message.get("URL", doi and f"https://doi.org/{doi}" or "")
    creators = []
    for author in message.get("author", []):
        given = author.get("given", "").strip()
        family = author.get("family", "").strip()
        name = " ".join(part for part in [given, family] if part)
        if name:
            creators.append(name)
    authors = " and ".join(creators)

    year = ""
    for key in ("published-print", "published-online", "created", "issued"):
        parts = message.get(key, {}).get("date-parts", [])
        if parts and parts[0]:
            year = str(parts[0][0])
            break

    citekey = make_datacite_citekey(title, creators, year or "nd")
    entry = {
        "ENTRYTYPE": "article",
        "ID": citekey or f"row{row_number}",
        "title": title,
        "author": authors,
        "year": year,
        "journal": (message.get("container-title") or [""])[0],
        "doi": doi,
        "url": url,
    }
    return {key: value for key, value in entry.items() if value}


def minimal_doi_bib_entry(row: ReviewedRow, doi: str) -> dict[str, str]:
    title = row.reference_text or row.identifier or f"reviewed-item-{row.row_number}"
    citekey = make_datacite_citekey(title, [], "nd")
    return {
        "ENTRYTYPE": "misc",
        "ID": citekey or f"row{row.row_number}",
        "title": title,
        "doi": doi,
        "url": f"https://doi.org/{doi}",
        "note": row.review_note,
    }


def build_webpage_entry(row: ReviewedRow) -> dict[str, str]:
    url = row.identifier if row.identifier.lower().startswith(("http://", "https://")) else ""
    title = row.reference_text or url or f"reviewed-item-{row.row_number}"
    key_root = re.sub(r"[^a-z0-9]+", "", title.lower())[:24] or f"row{row.row_number}"
    return {
        "ENTRYTYPE": "misc",
        "ID": f"{key_root}{row.row_number}",
        "title": title,
        "url": url,
        "note": row.review_note,
    }


def get_clients() -> tuple[Any | None, Any | None]:
    try:
        from pyzotero import zotero as pyzotero_client
    except ModuleNotFoundError:
        return None, None

    read_client = pyzotero_client.Zotero("0", "user", local=True)

    api_key = os.getenv("ZOTERO_API_KEY")
    web_library_id = os.getenv("ZOTERO_LIBRARY_ID")
    web_library_type = os.getenv("ZOTERO_LIBRARY_TYPE", "user")
    if api_key and web_library_id:
        write_client = pyzotero_client.Zotero(web_library_id, web_library_type, api_key=api_key, local=False)
        return read_client, write_client

    return read_client, None


def extract_created_key(result: Any) -> str:
    if isinstance(result, dict):
        success = result.get("success") or {}
        if success:
            return str(next(iter(success.values())))
    if hasattr(result, "json"):
        try:
            payload = result.json()
        except Exception:
            payload = None
        if isinstance(payload, dict):
            success = payload.get("success") or {}
            if success:
                return str(next(iter(success.values())))
    return ""


def find_existing_item(
    read_client: Any | None,
    *,
    row: ReviewedRow,
    existing_cache: dict[str, str],
    doi: str = "",
    url: str = "",
) -> str:
    if read_client is None:
        return ""

    if doi:
        cache_key = f"doi:{doi.lower()}"
        if cache_key in existing_cache:
            return existing_cache[cache_key]
        try:
            items = read_client.items(q=doi, limit=5)
        except Exception:
            items = []
        for item in items:
            data = item.get("data", {})
            if data.get("DOI", "").lower() == doi.lower():
                key = item.get("key", "")
                existing_cache[cache_key] = key
                return key
        existing_cache[cache_key] = ""

    if url:
        cache_key = f"url:{url}"
        if cache_key in existing_cache:
            return existing_cache[cache_key]
        try:
            items = read_client.items(q=url, limit=5)
        except Exception:
            items = []
        for item in items:
            data = item.get("data", {})
            if data.get("url", "") == url:
                key = item.get("key", "")
                existing_cache[cache_key] = key
                return key
        existing_cache[cache_key] = ""

    if row.identifier and not doi and not url:
        cache_key = f"id:{row.identifier}"
        if cache_key in existing_cache:
            return existing_cache[cache_key]
        try:
            items = read_client.items(q=row.identifier, limit=5)
        except Exception:
            items = []
        for item in items:
            key = item.get("key", "")
            if key:
                existing_cache[cache_key] = key
                return key
        existing_cache[cache_key] = ""

    return ""


def create_web_item(write_client: Any, row: ReviewedRow, collection_key: str, tags: list[str]) -> str:
    item = {
        "itemType": "webpage",
        "title": row.reference_text or row.identifier or f"reviewed-item-{row.row_number}",
        "url": row.identifier,
        "collections": [collection_key],
        "tags": [{"tag": tag} for tag in tags],
        "extra": row.review_note,
    }
    result = write_client.create_items([item])
    created_key = extract_created_key(result)
    if created_key:
        return created_key
    raise RuntimeError(f"Failed to create webpage item: {result}")


def create_doi_item(
    write_client: Any,
    row: ReviewedRow,
    message: dict[str, Any] | None,
    doi: str,
    collection_key: str,
    tags: list[str],
) -> str:
    payload = {
        "itemType": "journalArticle",
        "title": row.reference_text or row.identifier or f"reviewed-item-{row.row_number}",
        "DOI": doi,
        "url": f"https://doi.org/{doi}" if doi else "",
        "collections": [collection_key],
        "tags": [{"tag": tag} for tag in tags],
        "extra": row.review_note,
    }

    message = message or {}
    titles = message.get("title") or []
    if titles:
        payload["title"] = titles[0]
    if message.get("URL"):
        payload["url"] = message["URL"]

    creators = []
    for author in message.get("author", []):
        family = author.get("family", "").strip()
        given = author.get("given", "").strip()
        if family:
            creators.append({"creatorType": "author", "firstName": given, "lastName": family})
    if creators:
        payload["creators"] = creators

    for key in ("published-print", "published-online", "created", "issued"):
        parts = message.get(key, {}).get("date-parts", [])
        if parts and parts[0]:
            payload["date"] = "-".join(str(part) for part in parts[0])
            break

    container = (message.get("container-title") or [""])[0]
    if container:
        payload["publicationTitle"] = container

    result = write_client.create_items([payload])
    created_key = extract_created_key(result)
    if created_key:
        return created_key
    raise RuntimeError(f"Failed to create DOI item: {result}")


def probe_local_write(read_client: Any | None, *, collection_key: str, probe_path: Path) -> LocalWriteProbe:
    cached = load_json(probe_path)
    required = {"supported", "checked_at", "collection_key", "collection_found"}
    if cached and required.issubset(cached):
        return LocalWriteProbe(
            supported=bool(cached.get("supported")),
            checked_at=str(cached.get("checked_at", "")),
            collection_key=str(cached.get("collection_key", collection_key)),
            collection_found=bool(cached.get("collection_found")),
            created_key=str(cached.get("created_key", "")),
            cleaned_up=bool(cached.get("cleaned_up", False)),
            reason=str(cached.get("reason", "")),
            raw_result=str(cached.get("raw_result", "")),
        )

    checked_at = datetime.now().isoformat(timespec="seconds")
    if read_client is None:
        probe = LocalWriteProbe(
            supported=False,
            checked_at=checked_at,
            collection_key=collection_key,
            collection_found=False,
            reason="pyzotero is not installed, so local Zotero write cannot be tested.",
        )
        write_json_atomic(probe_path, asdict(probe))
        return probe

    try:
        collection = read_client.collection(collection_key)
        collection_found = collection.get("key", "") == collection_key
    except Exception as exc:
        probe = LocalWriteProbe(
            supported=False,
            checked_at=checked_at,
            collection_key=collection_key,
            collection_found=False,
            reason=f"Could not read target collection before probing local write: {describe_exception(exc)}",
        )
        write_json_atomic(probe_path, asdict(probe))
        return probe

    probe_title = f"Codex local write probe {checked_at}"
    payload = [
        {
            "itemType": "webpage",
            "title": probe_title,
            "url": "https://example.invalid/zotero-local-write-probe",
            "collections": [collection_key],
            "tags": [{"tag": "workflow/local-write-probe"}],
            "extra": "Temporary probe item; safe to delete.",
        }
    ]

    created_key = ""
    raw_result = ""
    try:
        result = read_client.create_items(payload)
        raw_result = json.dumps(result, ensure_ascii=False) if isinstance(result, (dict, list)) else repr(result)
        created_key = extract_created_key(result)
    except Exception as exc:
        probe = LocalWriteProbe(
            supported=False,
            checked_at=checked_at,
            collection_key=collection_key,
            collection_found=collection_found,
            reason=f"Local pyzotero create_items probe failed: {describe_exception(exc)}",
        )
        write_json_atomic(probe_path, asdict(probe))
        return probe

    if not created_key:
        probe = LocalWriteProbe(
            supported=False,
            checked_at=checked_at,
            collection_key=collection_key,
            collection_found=collection_found,
            reason="Local pyzotero create_items returned no created item key.",
            raw_result=raw_result,
        )
        write_json_atomic(probe_path, asdict(probe))
        return probe

    try:
        created_item = read_client.item(created_key)
        delete_response = read_client.delete_item([created_item])
        delete_ok = getattr(delete_response, "status_code", None) in {200, 204}
    except Exception as exc:
        probe = LocalWriteProbe(
            supported=False,
            checked_at=checked_at,
            collection_key=collection_key,
            collection_found=collection_found,
            created_key=created_key,
            cleaned_up=False,
            reason=f"Local write probe created item {created_key}, but cleanup failed: {describe_exception(exc)}",
            raw_result=raw_result,
        )
        write_json_atomic(probe_path, asdict(probe))
        return probe

    probe = LocalWriteProbe(
        supported=delete_ok,
        checked_at=checked_at,
        collection_key=collection_key,
        collection_found=collection_found,
        created_key=created_key,
        cleaned_up=delete_ok,
        reason="Local pyzotero write probe created and removed a temporary item successfully."
        if delete_ok
        else f"Local write probe created item {created_key}, but delete response was {getattr(delete_response, 'status_code', None)}.",
        raw_result=raw_result,
    )
    write_json_atomic(probe_path, asdict(probe))
    return probe


def save_checkpoint(
    checkpoint_path: Path,
    *,
    input_path: Path,
    input_sha1: str,
    run_dir: Path,
    mode_requested: str,
    mode_used: str,
    collection_key: str,
    processed_row_numbers: set[int],
    results: list[ImportResult],
    bib_entries: list[dict[str, Any]],
    stop_reason: str = "",
) -> None:
    payload = {
        "input_path": str(input_path),
        "input_sha1": input_sha1,
        "run_dir": str(run_dir),
        "mode_requested": mode_requested,
        "mode_used": mode_used,
        "collection_key": collection_key,
        "processed_row_numbers": sorted(processed_row_numbers),
        "results": [asdict(item) for item in results],
        "bib_entries": bib_entries,
        "stop_reason": stop_reason,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    write_json_atomic(checkpoint_path, payload)


def load_checkpoint(checkpoint_path: Path) -> tuple[dict[str, Any], set[int], list[ImportResult], list[dict[str, Any]], str]:
    payload = load_json(checkpoint_path)
    processed = set(payload.get("processed_row_numbers", []))
    results = [ImportResult(**item) for item in payload.get("results", [])]
    bib_entries = payload.get("bib_entries", [])
    stop_reason = payload.get("stop_reason", "")
    return payload, processed, results, bib_entries, stop_reason


def save_cache(cache_file: Path, metadata_cache: dict[str, dict[str, Any] | None], existing_cache: dict[str, str]) -> None:
    write_json_atomic(
        cache_file,
        {
            "metadata_cache": metadata_cache,
            "existing_cache": existing_cache,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        },
    )


def write_run_outputs(run_dir: Path, rows: list[ReviewedRow], results: list[ImportResult], bib_entries: list[dict[str, Any]]) -> None:
    (run_dir / "reviewed_rows.json").write_text(
        json.dumps([asdict(row) for row in rows], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (run_dir / "import_results.json").write_text(
        json.dumps([asdict(result) for result in results], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    fieldnames = list(ImportResult.__annotations__.keys())
    with (run_dir / "import_results.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for result in results:
            writer.writerow(asdict(result))

    lines = ["# Import Report", ""]
    for result in results:
        lines.extend(
            [
                f"## Row {result.row_number} | {result.action}",
                f"- Identifier: {result.identifier or 'N/A'}",
                f"- Source type: {result.source_type}",
                f"- Title: {result.title or 'N/A'}",
                f"- DOI: {result.doi or 'N/A'}",
                f"- URL: {result.url or 'N/A'}",
                f"- Zotero key: {result.zotero_key or 'N/A'}",
                f"- Reason: {result.reason or 'N/A'}",
                "",
            ]
        )
    (run_dir / "import_report.md").write_text("\n".join(lines), encoding="utf-8")

    db = bibtexparser.bibdatabase.BibDatabase()
    deduped_entries: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for entry in bib_entries:
        entry_id = entry.get("ID", "")
        if entry_id and entry_id in seen_ids:
            continue
        if entry_id:
            seen_ids.add(entry_id)
        deduped_entries.append(entry)
    db.entries = deduped_entries
    writer = BibTexWriter()
    writer.order_entries_by = None
    (run_dir / "verified_from_review.bib").write_text(writer.write(db), encoding="utf-8")
    (run_dir / "verified_dois.txt").write_text(
        "\n".join(entry["doi"] for entry in deduped_entries if entry.get("doi")),
        encoding="utf-8",
    )


def update_resume_state(
    state_dir: Path,
    *,
    source_path: Path,
    run_dir: Path,
    mode_requested: str,
    mode_used: str,
    collection_name: str,
    collection_key: str,
    total_rows: int,
    processed_rows: int,
    results: list[ImportResult],
    blockers: list[str],
    stop_reason: str = "",
    backend_probe: LocalWriteProbe | None = None,
) -> None:
    counts = {
        "total_rows": total_rows,
        "processed_rows": processed_rows,
        "imported": sum(1 for item in results if item.action == "imported"),
        "prepared_only": sum(1 for item in results if item.action == "prepared_only"),
        "duplicate": sum(1 for item in results if item.action == "duplicate"),
        "blocked": sum(1 for item in results if item.action == "blocked"),
    }

    if blockers:
        if backend_probe and not backend_probe.supported:
            next_step = "Keep the prepared outputs, or configure Zotero Web API credentials and rerun with --mode hybrid."
        else:
            next_step = "Resolve the blocker, then rerun the same command with the same --run-dir to continue."
    elif stop_reason == "runtime_limit":
        next_step = "Rerun the same command with the same --run-dir to resume from checkpoint."
    elif stop_reason == "row_limit":
        next_step = "Increase --max-rows or remove it, then rerun with the same --run-dir to continue."
    elif stop_reason == "interrupted":
        next_step = "Rerun the same command with the same --run-dir to continue from checkpoint."
    else:
        next_step = "Open the latest run artifacts and verify the prepared/imported items for the cpu collection."

    status = "ready"
    if blockers:
        status = "blocked"
    elif stop_reason and processed_rows < total_rows:
        status = "partial"

    payload: dict[str, Any] = {
        "status": status,
        "mode_requested": mode_requested,
        "mode_used": mode_used,
        "input_path": str(source_path),
        "collection_name": collection_name,
        "collection_key": collection_key,
        "run_dir": str(run_dir),
        "counts": counts,
        "blockers": blockers,
        "stop_reason": stop_reason,
        "next_step": next_step,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    if backend_probe is not None:
        payload["backend_probe"] = asdict(backend_probe)
    write_json_atomic(state_dir / "latest_run.json", payload)

    resume_lines = [
        "# Resume State",
        "",
        f"- Source: `{source_path}`",
        f"- Target collection: `{collection_name}` (`{collection_key}`)",
        f"- Requested mode: `{mode_requested}`",
        f"- Actual mode: `{mode_used}`",
        f"- Run directory: `{run_dir}`",
        f"- Processed rows: {processed_rows}/{total_rows}",
        f"- Imported: {counts['imported']}",
        f"- Prepared only: {counts['prepared_only']}",
        f"- Duplicates: {counts['duplicate']}",
        f"- Blocked: {counts['blocked']}",
        f"- Stop reason: {stop_reason or 'none'}",
    ]
    if backend_probe is not None:
        resume_lines.extend(
            [
                f"- Local probe supported: {backend_probe.supported}",
                f"- Local probe reason: {backend_probe.reason or 'n/a'}",
            ]
        )
    resume_lines.extend(["", "## Blockers"])
    if blockers:
        resume_lines.extend([f"- {blocker}" for blocker in blockers])
    else:
        resume_lines.append("- none")
    resume_lines.extend(["", "## Next Step", f"- {next_step}", ""])
    (state_dir / "RESUME.md").write_text("\n".join(resume_lines), encoding="utf-8")


def sync_progress(
    *,
    checkpoint_path: Path,
    cache_path: Path,
    state_dir: Path,
    input_path: Path,
    input_sha1: str,
    run_dir: Path,
    mode_requested: str,
    mode_used: str,
    collection_name: str,
    collection_key: str,
    rows: list[ReviewedRow],
    processed_row_numbers: set[int],
    results: list[ImportResult],
    bib_entries: list[dict[str, Any]],
    metadata_cache: dict[str, dict[str, Any] | None],
    existing_cache: dict[str, str],
    blockers: list[str],
    stop_reason: str,
    backend_probe: LocalWriteProbe | None,
) -> None:
    save_checkpoint(
        checkpoint_path,
        input_path=input_path,
        input_sha1=input_sha1,
        run_dir=run_dir,
        mode_requested=mode_requested,
        mode_used=mode_used,
        collection_key=collection_key,
        processed_row_numbers=processed_row_numbers,
        results=results,
        bib_entries=bib_entries,
        stop_reason=stop_reason,
    )
    save_cache(cache_path, metadata_cache, existing_cache)
    write_run_outputs(run_dir, rows, results, bib_entries)
    update_resume_state(
        state_dir,
        source_path=input_path,
        run_dir=run_dir,
        mode_requested=mode_requested,
        mode_used=mode_used,
        collection_name=collection_name,
        collection_key=collection_key,
        total_rows=len(rows),
        processed_rows=len(processed_row_numbers),
        results=results,
        blockers=blockers,
        stop_reason=stop_reason,
        backend_probe=backend_probe,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Import or prepare reviewed bibliography rows from an XLSX sheet for Zotero.")
    parser.add_argument("--input", required=True, help="Path to the reviewed XLSX file")
    parser.add_argument("--sheet", help="Optional sheet name; defaults to the first sheet")
    parser.add_argument("--collection-name", default=DEFAULT_COLLECTION_NAME)
    parser.add_argument("--collection-key", default=DEFAULT_COLLECTION_KEY)
    parser.add_argument("--mode", choices=["auto", "prepare", "local", "hybrid"], default="auto")
    parser.add_argument("--run-dir", help="Optional explicit run directory")
    parser.add_argument("--checkpoint-every", type=int, default=DEFAULT_CHECKPOINT_EVERY)
    parser.add_argument("--max-rows", type=int, help="Process at most this many new rows in this invocation")
    parser.add_argument("--stop-after-seconds", type=int, help="Stop after roughly this many seconds and leave a resumable checkpoint")
    parser.add_argument("--probe-only", action="store_true", help="Only probe local pyzotero write support and write resumable state")
    parser.add_argument("--force-restart", action="store_true", help="Ignore any existing checkpoint in the run directory and start fresh")
    parser.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT)
    parser.add_argument("--retries", type=int, default=DEFAULT_RETRIES)
    args = parser.parse_args()

    source_path = Path(args.input).expanduser().resolve()
    if not source_path.exists():
        raise FileNotFoundError(f"Input file not found: {source_path}")

    project_root = Path(__file__).resolve().parent.parent
    runs_root = project_root / "runs"
    state_dir = project_root / "state"
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = Path(args.run_dir).expanduser().resolve() if args.run_dir else runs_root / stamp
    run_dir.mkdir(parents=True, exist_ok=True)

    copied_input = run_dir / source_path.name
    if not copied_input.exists():
        shutil.copy2(source_path, copied_input)

    checkpoint_path = run_dir / "checkpoint.json"
    cache_path = run_dir / "cache.json"
    probe_path = run_dir / "local_write_probe.json"
    input_sha1 = sha1_file(source_path)

    rows = load_reviewed_rows(source_path, sheet_name=args.sheet)
    read_client, hybrid_client = get_clients()
    local_probe = probe_local_write(read_client, collection_key=args.collection_key, probe_path=probe_path)

    blockers: list[str] = []
    write_client: Any | None = None
    mode_used = args.mode
    if args.mode == "auto":
        if local_probe.supported:
            mode_used = "local"
            write_client = read_client
        else:
            mode_used = "prepare"
            blockers.append(local_probe.reason)
    elif args.mode == "local":
        if local_probe.supported:
            mode_used = "local"
            write_client = read_client
        else:
            mode_used = "prepare"
            blockers.append(local_probe.reason)
    elif args.mode == "hybrid":
        if hybrid_client is not None:
            mode_used = "hybrid"
            write_client = hybrid_client
        else:
            mode_used = "prepare"
            blockers.append("Hybrid mode requested, but ZOTERO_API_KEY and ZOTERO_LIBRARY_ID are not configured.")
    else:
        mode_used = "prepare"

    metadata_cache: dict[str, dict[str, Any] | None] = {}
    existing_cache: dict[str, str] = {}
    cached_state = load_json(cache_path)
    if cached_state:
        metadata_cache = cached_state.get("metadata_cache", {}) or {}
        existing_cache = cached_state.get("existing_cache", {}) or {}

    processed_row_numbers: set[int] = set()
    results: list[ImportResult] = []
    bib_entries: list[dict[str, Any]] = []
    if checkpoint_path.exists() and not args.force_restart:
        checkpoint_payload, processed_row_numbers, results, bib_entries, _ = load_checkpoint(checkpoint_path)
        if checkpoint_payload.get("input_sha1") and checkpoint_payload.get("input_sha1") != input_sha1:
            raise RuntimeError("Existing checkpoint belongs to a different input file. Use --force-restart or a new --run-dir.")

    if args.probe_only:
        sync_progress(
            checkpoint_path=checkpoint_path,
            cache_path=cache_path,
            state_dir=state_dir,
            input_path=source_path,
            input_sha1=input_sha1,
            run_dir=run_dir,
            mode_requested=args.mode,
            mode_used=mode_used,
            collection_name=args.collection_name,
            collection_key=args.collection_key,
            rows=rows,
            processed_row_numbers=processed_row_numbers,
            results=results,
            bib_entries=bib_entries,
            metadata_cache=metadata_cache,
            existing_cache=existing_cache,
            blockers=blockers,
            stop_reason="probe_only",
            backend_probe=local_probe,
        )
        print(f"Probe supported: {local_probe.supported}")
        print(f"Probe reason: {local_probe.reason}")
        print(f"Run directory: {run_dir}")
        return

    cr = Crossref()
    tags = [
        "status/verified",
        "source/reviewed-xlsx",
        "workflow/zotero-reviewed-import",
    ]

    processed_this_run = 0
    stop_reason = ""
    started_at = time.monotonic()

    def flush(current_stop_reason: str) -> None:
        sync_progress(
            checkpoint_path=checkpoint_path,
            cache_path=cache_path,
            state_dir=state_dir,
            input_path=source_path,
            input_sha1=input_sha1,
            run_dir=run_dir,
            mode_requested=args.mode,
            mode_used=mode_used,
            collection_name=args.collection_name,
            collection_key=args.collection_key,
            rows=rows,
            processed_row_numbers=processed_row_numbers,
            results=results,
            bib_entries=bib_entries,
            metadata_cache=metadata_cache,
            existing_cache=existing_cache,
            blockers=blockers,
            stop_reason=current_stop_reason,
            backend_probe=local_probe,
        )

    try:
        for row in rows:
            if row.row_number in processed_row_numbers:
                continue
            if args.max_rows is not None and processed_this_run >= args.max_rows:
                stop_reason = "row_limit"
                break
            if args.stop_after_seconds is not None and time.monotonic() - started_at >= args.stop_after_seconds:
                stop_reason = "runtime_limit"
                break

            source_type = normalize_source_type(row.identifier)
            try:
                if source_type == "doi":
                    message, doi = fetch_crossref_payload(
                        row.identifier,
                        row.reference_text,
                        cr,
                        metadata_cache=metadata_cache,
                        timeout=args.timeout,
                        retries=args.retries,
                    )
                    if doi:
                        bib_entries.append(payload_to_bib_entry(message, row.reference_text, row.row_number) if message else minimal_doi_bib_entry(row, doi))
                    else:
                        bib_entries.append(minimal_doi_bib_entry(row, doi))

                    if not doi:
                        results.append(
                            ImportResult(
                                row_number=row.row_number,
                                identifier=row.identifier,
                                reference_text=row.reference_text,
                                source_type=source_type,
                                action="blocked",
                                reason="Could not resolve DOI from the reviewed row.",
                            )
                        )
                    else:
                        url = (message or {}).get("URL", f"https://doi.org/{doi}")
                        title = ((message or {}).get("title") or [row.reference_text or row.identifier or ""])[0]
                        existing_key = find_existing_item(read_client, row=row, existing_cache=existing_cache, doi=doi, url=url)
                        if existing_key:
                            results.append(
                                ImportResult(
                                    row_number=row.row_number,
                                    identifier=row.identifier,
                                    reference_text=row.reference_text,
                                    source_type=source_type,
                                    action="duplicate",
                                    zotero_key=existing_key,
                                    title=title,
                                    doi=doi,
                                    url=url,
                                    reason="Matching Zotero item already exists.",
                                )
                            )
                        elif write_client is not None and mode_used in {"local", "hybrid"}:
                            zotero_key = create_doi_item(write_client, row, message, doi, args.collection_key, tags)
                            results.append(
                                ImportResult(
                                    row_number=row.row_number,
                                    identifier=row.identifier,
                                    reference_text=row.reference_text,
                                    source_type=source_type,
                                    action="imported",
                                    zotero_key=zotero_key,
                                    title=title,
                                    doi=doi,
                                    url=url,
                                    reason="Imported into Zotero.",
                                )
                            )
                        else:
                            results.append(
                                ImportResult(
                                    row_number=row.row_number,
                                    identifier=row.identifier,
                                    reference_text=row.reference_text,
                                    source_type=source_type,
                                    action="prepared_only",
                                    title=title,
                                    doi=doi,
                                    url=url,
                                    reason="Prepared from DOI metadata; no write backend used.",
                                )
                            )

                elif source_type == "url":
                    bib_entries.append(build_webpage_entry(row))
                    existing_key = find_existing_item(read_client, row=row, existing_cache=existing_cache, url=row.identifier)
                    if existing_key:
                        results.append(
                            ImportResult(
                                row_number=row.row_number,
                                identifier=row.identifier,
                                reference_text=row.reference_text,
                                source_type=source_type,
                                action="duplicate",
                                zotero_key=existing_key,
                                title=row.reference_text,
                                url=row.identifier,
                                reason="Matching Zotero URL item already exists.",
                            )
                        )
                    elif write_client is not None and mode_used in {"local", "hybrid"}:
                        zotero_key = create_web_item(write_client, row, args.collection_key, tags)
                        results.append(
                            ImportResult(
                                row_number=row.row_number,
                                identifier=row.identifier,
                                reference_text=row.reference_text,
                                source_type=source_type,
                                action="imported",
                                zotero_key=zotero_key,
                                title=row.reference_text,
                                url=row.identifier,
                                reason="Imported as webpage item.",
                            )
                        )
                    else:
                        results.append(
                            ImportResult(
                                row_number=row.row_number,
                                identifier=row.identifier,
                                reference_text=row.reference_text,
                                source_type=source_type,
                                action="prepared_only",
                                title=row.reference_text,
                                url=row.identifier,
                                reason="Prepared from reviewed URL row; no write backend used.",
                            )
                        )

                else:
                    results.append(
                        ImportResult(
                            row_number=row.row_number,
                            identifier=row.identifier,
                            reference_text=row.reference_text,
                            source_type=source_type,
                            action="blocked",
                            reason="Row has no DOI or URL identifier.",
                        )
                    )
            except Exception as exc:
                results.append(
                    ImportResult(
                        row_number=row.row_number,
                        identifier=row.identifier,
                        reference_text=row.reference_text,
                        source_type=source_type,
                        action="blocked",
                        reason=describe_exception(exc),
                    )
                )

            processed_row_numbers.add(row.row_number)
            processed_this_run += 1
            if args.checkpoint_every > 0 and processed_this_run % args.checkpoint_every == 0:
                flush("")

    except KeyboardInterrupt:
        stop_reason = "interrupted"
    finally:
        flush(stop_reason)

    print(f"Total verified rows: {len(rows)}")
    print(f"Processed rows overall: {len(processed_row_numbers)}")
    print(f"Processed rows this run: {processed_this_run}")
    print(f"Mode used: {mode_used}")
    print(f"Run directory: {run_dir}")
    print(f"Local probe supported: {local_probe.supported}")
    print(f"Local probe reason: {local_probe.reason}")
    if blockers:
        print("Blockers:")
        for blocker in blockers:
            print(f"- {blocker}")
    if stop_reason:
        print(f"Stop reason: {stop_reason}")


if __name__ == "__main__":
    main()
